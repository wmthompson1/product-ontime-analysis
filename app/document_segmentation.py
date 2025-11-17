"""
Document Segmentation Module
=============================
Segments Excel documents into blocks based on cell ranges and segment types.
Perfect for hybrid RAG: unstructured metadata + structured tables.

Author: Manufacturing Intelligence Team
Purpose: Berkeley Haas AI Strategy - Advanced RAG for Invoice Processing
"""

import pandas as pd
import openpyxl
from io import BytesIO, StringIO
from typing import Dict, List, Tuple, Optional


def parse_segmentation_scheme(scheme_content: str) -> pd.DataFrame:
    """
    Parse segmentation scheme from CSV content.
    
    Args:
        scheme_content: CSV string with columns: Doc, block, upper_left, lower_right, Segment type
        
    Returns:
        DataFrame with segmentation scheme
    """
    df = pd.read_csv(StringIO(scheme_content))
    
    required_cols = ['Doc', 'block', 'upper_left', 'lower_right', 'Segment type']
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Missing required column: {col}")
    
    return df


def excel_range_to_indices(range_str: str, sheet) -> Tuple[int, int, int, int]:
    """
    Convert Excel range (e.g., 'A3:B5') to 0-indexed row/col indices.
    
    Args:
        range_str: Excel range like 'A3:B5' or 'A8:Doc 1 end'
        sheet: openpyxl worksheet object
        
    Returns:
        Tuple of (start_row, end_row, start_col, end_col) - 0-indexed
    """
    if ':' not in range_str:
        raise ValueError(f"Invalid range format: {range_str}")
    
    parts = range_str.split(':')
    upper_left = parts[0].strip()
    lower_right = parts[1].strip()
    
    if 'end' in lower_right.lower():
        cell = openpyxl.utils.cell.coordinate_from_string(upper_left)
        col_letter = cell[0]
        start_row = cell[1]
        
        start_col = openpyxl.utils.column_index_from_string(col_letter)
        end_row = sheet.max_row
        end_col = sheet.max_column
        
        return (start_row - 1, end_row - 1, start_col - 1, end_col - 1)
    else:
        upper_cell = openpyxl.utils.cell.coordinate_from_string(upper_left)
        lower_cell = openpyxl.utils.cell.coordinate_from_string(lower_right)
        
        start_row = upper_cell[1] - 1
        end_row = lower_cell[1] - 1
        start_col = openpyxl.utils.column_index_from_string(upper_cell[0]) - 1
        end_col = openpyxl.utils.column_index_from_string(lower_cell[0]) - 1
        
        return (start_row, end_row, start_col, end_col)


def extract_freeform_block(sheet, start_row: int, end_row: int, start_col: int, end_col: int) -> Dict:
    """
    Extract free-form block as key-value pairs.
    
    Args:
        sheet: openpyxl worksheet
        start_row, end_row, start_col, end_col: 0-indexed cell range
        
    Returns:
        Dictionary of key-value pairs
    """
    metadata = {}
    
    for row_idx in range(start_row, end_row + 1):
        key_cell = sheet.cell(row=row_idx + 1, column=start_col + 1)
        value_cell = sheet.cell(row=row_idx + 1, column=start_col + 2) if end_col >= start_col + 1 else None
        
        key = key_cell.value
        value = value_cell.value if value_cell else None
        
        if key is not None and str(key).strip():
            key_clean = str(key).strip()
            metadata[key_clean] = value
    
    return metadata


def extract_tabular_block(sheet, start_row: int, end_row: int, start_col: int, end_col: int) -> Dict:
    """
    Extract tabular block as DataFrame structure.
    
    Args:
        sheet: openpyxl worksheet
        start_row, end_row, start_col, end_col: 0-indexed cell range
        
    Returns:
        Dictionary with 'columns' and 'data' keys
    """
    data = []
    
    for row_idx in range(start_row, end_row + 1):
        row_data = []
        for col_idx in range(start_col, end_col + 1):
            cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)
            row_data.append(cell.value)
        data.append(row_data)
    
    if not data:
        return {'columns': [], 'data': []}
    
    columns = data[0]
    rows = data[1:]
    
    result = {
        'columns': [str(c) if c is not None else f'col_{i}' for i, c in enumerate(columns)],
        'data': rows,
        'row_count': len(rows),
        'column_count': len(columns)
    }
    
    return result


def segment_document(excel_file, segmentation_scheme: Optional[str] = None) -> Dict:
    """
    Segment Excel document based on segmentation scheme.
    
    Args:
        excel_file: File-like object or path to Excel file
        segmentation_scheme: Optional CSV string with segmentation scheme.
                            If None, uses default scheme for invoice_extract_sample1.xlsx
        
    Returns:
        Dictionary with segmented blocks and metadata
    """
    default_scheme = """Doc,block,upper_left,lower_right,Segment type
1,1,A3,B5,Free-form
1,2,A8,Doc 1 end,Tabular-form"""
    
    scheme = segmentation_scheme if segmentation_scheme else default_scheme
    scheme_df = parse_segmentation_scheme(scheme)
    
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    
    result = {
        'document_id': int(scheme_df['Doc'].iloc[0]),
        'total_blocks': len(scheme_df),
        'blocks': []
    }
    
    for idx, row in scheme_df.iterrows():
        doc_id = int(row['Doc'])
        block_id = int(row['block'])
        range_str = f"{row['upper_left']}:{row['lower_right']}"
        segment_type = row['Segment type']
        
        start_row, end_row, start_col, end_col = excel_range_to_indices(range_str, sheet)
        
        block_data = {
            'block_id': block_id,
            'range': range_str,
            'segment_type': segment_type,
            'cells': {
                'start_row': start_row + 1,
                'end_row': end_row + 1,
                'start_col': start_col + 1,
                'end_col': end_col + 1
            }
        }
        
        if segment_type.lower() == 'free-form':
            block_data['content'] = extract_freeform_block(sheet, start_row, end_row, start_col, end_col)
            block_data['content_type'] = 'metadata'
        elif segment_type.lower() == 'tabular-form':
            block_data['content'] = extract_tabular_block(sheet, start_row, end_row, start_col, end_col)
            block_data['content_type'] = 'table'
        else:
            block_data['content'] = None
            block_data['content_type'] = 'unknown'
        
        result['blocks'].append(block_data)
    
    wb.close()
    
    return result


def format_segmentation_report(segmentation_result: Dict) -> str:
    """
    Format segmentation result as human-readable report.
    
    Args:
        segmentation_result: Output from segment_document()
        
    Returns:
        Formatted text report
    """
    lines = []
    lines.append("=" * 70)
    lines.append("DOCUMENT SEGMENTATION REPORT")
    lines.append("=" * 70)
    lines.append(f"\nDocument ID: {segmentation_result['document_id']}")
    lines.append(f"Total Blocks: {segmentation_result['total_blocks']}\n")
    
    for block in segmentation_result['blocks']:
        lines.append("-" * 70)
        lines.append(f"Block {block['block_id']}: {block['segment_type']}")
        lines.append(f"Range: {block['range']} (Rows {block['cells']['start_row']}-{block['cells']['end_row']}, Cols {block['cells']['start_col']}-{block['cells']['end_col']})")
        lines.append(f"Content Type: {block['content_type']}")
        
        if block['content_type'] == 'metadata':
            lines.append("\nMetadata:")
            for key, value in block['content'].items():
                lines.append(f"  • {key}: {value}")
        elif block['content_type'] == 'table':
            lines.append(f"\nTable Structure:")
            lines.append(f"  • Columns: {', '.join(block['content']['columns'])}")
            lines.append(f"  • Rows: {block['content']['row_count']}")
            lines.append(f"  • Columns: {block['content']['column_count']}")
            
            if block['content']['data']:
                lines.append(f"\n  First 5 rows:")
                for i, row in enumerate(block['content']['data'][:5]):
                    lines.append(f"    {i+1}. {row}")
        
        lines.append("")
    
    lines.append("=" * 70)
    return "\n".join(lines)
